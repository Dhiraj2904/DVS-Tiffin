from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.views.generic import CreateView
from .forms import RegisterForm
#from .models import RegistrationData
from django.contrib import messages
from django.shortcuts import get_list_or_404, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse
from .forms import UsersLoginForm, AdminLoginForm
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from RCF.core.forms import RegisterForm, ContactForm
from RCF.core.models import EmpInsert, EmpInsert2, carta, carorder, orderdisplay, OrderNow, menudisplay, otdisplay, ltdisplay, dtdisplay, alacartaorder
from django.views.generic import ListView
from django.template import RequestContext
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import get_template
from django.core.mail import EmailMessage
import openpyxl
from datetime import date
from django.template.loader import render_to_string

from django.contrib.auth.models import User


def save_all(request):
	if request.method=='POST':
	
		post=carorder()
		post.carid= request.POST.get('carid')
		post.carmenu= request.POST.get('carmenu')
		post.carqty= request.POST.get('carqty')
		post.caramt=request.POST.get('caramt')
		post.author=request.user
		post.save()
		messages.success(request, "Order placed Successfully!!")
		return render(request, 'index.html')  

	else:
		res1=alacartaorder.objects.all()
		return render(request, 'index.html', {'alacartaorder':res1})

	return render(request,'adminpage.html')





def Insertrecord(request):
	if request.method=='POST':
		username=None
		saverecord=OrderNow()
		saverecord.ordernow_id=request.POST.get('ordernow_id')
		saverecord.tiffin_type=request.POST.get('tiffin_type')
		saverecord.subtiffin_type=request.POST.get('subtiffin_type')
		saverecord.from_date=request.POST.get('from_date')
		saverecord.to_date=request.POST.get('to_date')
		saverecord.total_amount=request.POST.get('total_amount')
	
		#saverecord.author=User.objects.get(request.POST.get('username'))
		saverecord.author=request.user
		saverecord.save()
		
		#messages.success(request, "Order Placed Succesfully")
		return render(request, 'user.html')
	else:
		res1=carta.objects.all()
		return render(request, 'adminpage.html', {'carta':res1})


@login_required
def insertalacarta(request):
	if request.method=='POST':
		saverecord=alacartaorder()
		saverecord.aid=request.POST.get('aid')
		saverecord.amenu=request.POST.get('amenu')
		saverecord.aamount=request.POST.get('aamount')
		
		
		saverecord.save()
		
		#messages.success(request, 'Order Placed successfully!!..')
		return render(request, 'user.html')
	else:
		res1=carta.objects.all()
		res2=alacartaorder.objects.all()
		return render(request, 'adminpage.html', {'carta':res1, 'alacartaorder':res2})


def showsociety(request):
	results=showsoc.objects.all()
	return render(request, "signup.html", {"results":results})

def showmenu(request):
	r1=carta.objects.all()
	return render(request, "adminpage.html", {"carta":r1})

def showpastorder(request):
	res=OrderNow.objects.filter(author_id=request.user)
	return render(request, "user.html", {'OrderNow':res})


def showpastorder1(request):
	res=OrderNow.objects.all()
	
	return render(request, "base.html", {'OrderNow':res})
@login_required
def showmainmenu(request):
	result1=menudisplay.objects.all()
	result2=otdisplay.objects.all()
	result3=ltdisplay.objects.all()
	result4=dtdisplay.objects.all()
	return render(request, "menupage.html", {'menudisplay':result1, 'otdisplay':result2, 'ltdisplay':result3, 'dtdisplay':result4})
def offtiffdisplay(request):
	result1=otdisplay.objects.all()
	return render(request, "menupage.html", {'otdisplay':result1})

@login_required
def index(request):
    return render(request,'user.html')

def odnow(request):
	return render(request, 'ordernow.html')

def home(request):
	return render(request, 'home.html')


def signup(request):
	
	
	return render(request, 'registration/signup.html',context)





def register_view(request):
	form = RegisterForm(request.POST or None)
	#profile_form=UserProfileForm(request.POST or None)
	if form.is_valid():
		user = form.save()
		#profile=profile_form.save(commit=False)
		#profile.user=user
		#profile.save()
		
		username=form.cleaned_data.get("username")
		password = form.cleaned_data.get("password")	
		user.set_password(password)
		user.save()
		
		messages.success(request, 'Account Registered Successfully!!')
		form=RegisterForm()
		#profile_form=UserProfileForm()

	context = {"form":RegisterForm}
	return render(request, "registration/signup.html", {
		"title" : "Register",
		"form" : form,
	})


@login_required
def login_view(request):
	form = UsersLoginForm(request.POST or None)
	
	if form.is_valid():
		username = form.cleaned_data.get("username")
		password = form.cleaned_data.get("password")
		user = authenticate(username = username, password = password)
		if request.user.is_superuser:
			login(request, user)
			return render(request, 'support.html')
	else:
		return render(request, 'user.html')


@login_required
def adminlogin(request):
	form = AdminLoginForm(request.POST or None)
	
	if form.is_valid():
		username = form.cleaned_data.get("username")
		password = form.cleaned_data.get("password")
		user = authenticate(username = username, password = password)
		if request.user.is_superuser:
			login(request, user)
			#messages.messages.success("New account created")
			return HttpResponseRedirect("")
	else:
		res1=OrderNow.objects.all()
		return render(request, 'adminhome.html',{'OrderNow':res1})
		


@login_required
def user(request):
    return render(request, 'user.html')
def menupage(request):
    return render(request, 'menupage.html')
def adminpage(request):

    return render(request, 'adminpage.html')
def support(request):
	form = AdminLoginForm(request.POST or None)
	
	if form.is_valid():
		username = form.cleaned_data.get("username")
		password = form.cleaned_data.get("password")
		user = authenticate(username = username, password = password)
		if request.user.is_superuser:
			login(request, user)
			messages.messages.success("New account created")
			return render(request, 'demo.html')
	else:
    		return render(request, 'support.html')
def index(request):
    return render(request, 'index.html')


def showorder(request):

	return render_to_response('user.html', {'obj': models.OrderNow.objects.all()})

class PersonListView(ListView):
    model = EmpInsert2
    template_name = 'user.html'

def demo(request):
	result= OrderNow.objects.all()
	return render(request, 'adminhome.html', {'OrderNow':result})


def user_login(request):
	context = RequestContext(request)
	if request.method == 'POST':
		username = request.POST.get('username',False)
		password = request.POST.get('password',False)
		user = authenticate(username=username, password=password)
		if user is not None and user.is_superuser:
			login(request, user)
			return HttpResponseRedirect("myadmin")
	else:
		return render(request, 'adminhome.html')


def Contact(request):
	
	if request.method=='POST':
		name = request.POST.get('name')
		email = request.POST.get('email')
		message = request.POST.get('message')
		context= {'name':name, 'email':email, 'message':message}

		template=render_to_string('contact_form.txt',context)
		send_mail('Contact Form',
			template,
			settings.EMAIL_HOST_USER,
			['amitrajput111777@gmail.com'],
			fail_silently=False)

	return render(request, 'contact.html')



def readexcel(request):
	if "GET" == request.method:
		return render(request, 'adminhome.html')
	else:
		excel_file=request.FILES["excel_file"]

	#You may put the file extension check here
	wb=openpyxl.load_workbook(excel_file)

	# getting all sheets
	sheets = wb.sheetnames
	#print(sheets)

	#getting a particular sheet by name
	#Read School Worksheet
	schoolworksheet = wb["School Tiffin"]
	#print(schoolworksheet)

	#excel_data=list()
	#Iterate and get values from each cell
	menu_display=menudisplay()
	for rowno, rowval in enumerate(schoolworksheet.iter_rows(min_row=2, max_row=schoolworksheet.max_row),start=2):
		
		
		menu_display.mdate=schoolworksheet.cell(row=rowno,column=1).value
		menu_display.Vegetarian=(schoolworksheet.cell(row=rowno,column=2).value)
		menu_display.Mixed=(schoolworksheet.cell(row=rowno,column=3).value)
		menu_display.Premium_Veg=(schoolworksheet.cell(row=rowno,column=4).value)
		menu_display.Premium_Mixed=(schoolworksheet.cell(row=rowno,column=5).value)
		menu_display.save()


	officeworksheet = wb["Office Tiffin"]

	office_display=otdisplay()
	for rowno, rowval in enumerate(officeworksheet.iter_rows(min_row=2, max_row=officeworksheet.max_row),start=2):
		
		office_display.odate1=officeworksheet.cell(row=rowno,column=1).value
		office_display.Vegetarian1=(officeworksheet.cell(row=rowno,column=2).value)
		office_display.Mixed1=(officeworksheet.cell(row=rowno,column=3).value)
		office_display.Premium_Veg1=(officeworksheet.cell(row=rowno,column=4).value)
		office_display.Premium_Mixed1=(officeworksheet.cell(row=rowno,column=5).value)
		office_display.save()
	
	lunchworksheet = wb["Lunch Tiffin"]
	lunch_display=ltdisplay()
	for rowno, rowval in enumerate(lunchworksheet.iter_rows(min_row=2, max_row=lunchworksheet.max_row),start=2):
		
		lunch_display.ldate=lunchworksheet.cell(row=rowno,column=1).value
		lunch_display.Vegetarian=(lunchworksheet.cell(row=rowno,column=2).value)
		lunch_display.Mixed=(lunchworksheet.cell(row=rowno,column=3).value)
		lunch_display.Premium_Veg=(lunchworksheet.cell(row=rowno,column=4).value)
		lunch_display.Premium_Mixed=(lunchworksheet.cell(row=rowno,column=5).value)
		lunch_display.save()


	dinnerworksheet = wb["Dinner Tiffin"]
	dinner_display=dtdisplay()
	for rowno, rowval in enumerate(dinnerworksheet.iter_rows(min_row=2, max_row=dinnerworksheet.max_row),start=2):
		
		dinner_display.ddate=dinnerworksheet.cell(row=rowno,column=1).value
		dinner_display.Vegetarian=(dinnerworksheet.cell(row=rowno,column=2).value)
		dinner_display.Mixed=(dinnerworksheet.cell(row=rowno,column=3).value)
		dinner_display.Premium_Veg=(dinnerworksheet.cell(row=rowno,column=4).value)
		dinner_display.Premium_Mixed=(dinnerworksheet.cell(row=rowno,column=5).value)
		dinner_display.save()


	return render(request,'adminhome.html',{})


def readexcel2(request):
	if "GET" == request.method:
		return render(request, 'adminhome.html')
	else:
		excel_file2=request.FILES["excel_file2"]

	#You may put the file extension check here
	wb=openpyxl.load_workbook(excel_file2)

	# getting all sheets
	sheets = wb.sheetnames
	#print(sheets)

	#getting a particular sheet by name
	#Read School Worksheet
	schoolworksheet = wb["Sheet1"]
	#print(schoolworksheet)

	#excel_data=list()
	#Iterate and get values from each cell
	menu_display=alacartaorder()
	for rowno, rowval in enumerate(schoolworksheet.iter_rows(min_row=2, max_row=schoolworksheet.max_row),start=2):
		
		
		menu_display.amenu=schoolworksheet.cell(row=rowno,column=1).value
		menu_display.aamount=(schoolworksheet.cell(row=rowno,column=2).value)
		menu_display.save()


	return render(request,'adminhome.html',{})

def delrec(request, id):
	delrecord=OrderNow.objects.get(ordernow_id=id)
	delrecord.delete()

	res=OrderNow.objects.all()
	return render(request,'user.html',{'OrderNow':res})

